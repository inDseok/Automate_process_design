from typing import List
from backend.models import SubTree, SubNode
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from io import BytesIO
import re

EXCEL_PATH = r"C:\Users\USER\Desktop\공정설계 자동화\H_Lamp 공정설계 TOOL_korea Version 7.0\H_Lamp_공정설계 TOOL_Korea Version 7.0_230717.xlsm"
TARGET_SHEET = "STD_LHD_LD"


def list_sub_names() -> List[str]:
    return ["외주SUB(위트)"]


def get_cell_value(ws, r, c):
    coord = f"{get_column_letter(c)}{r}"

    for merged in ws.merged_cells.ranges:
        if coord in merged:
            return ws.cell(
                merged.min_row,
                merged.min_col
            ).value

    return ws.cell(row=r, column=c).value

def read_right_value(ws, r, c):
    col = c + 1
    max_col = ws.max_column
    visited_merged = set()

    while col <= max_col:
        coord = f"{get_column_letter(col)}{r}"

        value = ws.cell(row=r, column=col).value

        for merged in ws.merged_cells.ranges:
            if coord in merged:
                key = (merged.min_row, merged.min_col)
                if key in visited_merged:
                    value = None
                else:
                    visited_merged.add(key)
                    value = ws.cell(
                        merged.min_row,
                        merged.min_col
                    ).value
                break

        if value not in [None, "", "부품명", "품번", "수량", "재질"]:
            return str(value).strip()

        col += 1

    return None

def find_row_with_label(ws, start_r, start_c, label):
    for offset in range(1, 10):
        row = start_r + offset
        col = start_c
        cell = ws.cell(row=row, column=col)

        if cell.value:
            s = str(cell.value).replace(" ", "").replace("\n", "").strip()
            if label in s:
                return offset

        coord = f"{get_column_letter(col)}{row}"
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                top_left = ws.cell(merged.min_row, merged.min_col)
                if top_left.value:
                    s2 = str(top_left.value).replace(" ", "").replace("\n", "").strip()
                    if label in s2:
                        return offset

    raise ValueError(f"Label '{label}' not found below row {start_r}")


def read_part_no(ws, r):
    for col in range(1, 200):
        cell = ws.cell(row=r, column=col)
        if cell.value == "품번":
            val_cell = ws.cell(row=r, column=col + 1)
            return str(val_cell.value).strip() if val_cell.value is not None else None
    return None


def make_stable_id(sheet_name: str, row: int, col: int) -> str:
    return f"{sheet_name}:{row}:{col}"


def parse_block(ws, sheet_name: str, label_row: int, label_col: int):
    r = label_row
    c = label_col

    # 1) 부품명 (첫 줄)
    name = read_right_value(ws, r, c)

    # 2) 품번
    try:
        part_row = r + find_row_with_label(ws, r, c, "품번")
        part_no = read_right_value(ws, part_row, c)
    except Exception:
        part_no = None

    # 3) 수량
    try:
        qty_row = r + find_row_with_label(ws, r, c, "수량")
        raw_qty = read_right_value(ws, qty_row, c)
    except Exception:
        raw_qty = None

    # 4) 재질
    try:
        mat_row = r + find_row_with_label(ws, r, c, "재질")
        material = read_right_value(ws, mat_row, c)
    except Exception:
        material = None

    return {
        "id": make_stable_id(sheet_name, r, c),
        "name": name,
        "qty": raw_qty,
        "part_no": part_no,
        "material": material,
        "row": r,
        "col": c
    }




def build_tree_from_sheet(ws, sheet_name: str, sub_name: str) -> SubTree:
    boxes = []

    for row in ws.iter_rows(min_row=1, max_col=200, max_row=500):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if str(v).strip() == "부품명":
                box = parse_block(ws, sheet_name, cell.row, cell.column)
                boxes.append(box)

    boxes_sorted = sorted(boxes, key=lambda b: (b["row"], b["col"]))

    stack = []
    nodes: List[SubNode] = []

    for idx, box in enumerate(boxes_sorted):
        while stack and box["col"] <= stack[-1]["col"]:
            stack.pop()

        parent_id = stack[-1]["id"] if stack else None

        nodes.append(
            SubNode(
                id=box["id"],
                parent_id=parent_id,
                order=idx,
                type="PART",
                name=box["name"],
                part_no=box["part_no"],
                material=box["material"],
                qty=box["qty"]
            )
        )
        stack.append(box)

    return SubTree(sub_name=sub_name, nodes=nodes)


def load_sub_tree(sub_name: str) -> SubTree:
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb[TARGET_SHEET]
    except Exception as e:
        print("엑셀 로딩 에러:", e)
        raise

    return build_tree_from_sheet(ws, TARGET_SHEET, sub_name=sub_name)


def parse_uploaded_excel(binary_data: bytes) -> SubTree:
    wb = load_workbook(BytesIO(binary_data), data_only=True)

    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"{TARGET_SHEET} 시트를 찾을 수 없습니다.")

    ws = wb[TARGET_SHEET]

    # sub_name은 리스트와 동일하게 유지
    return build_tree_from_sheet(ws, TARGET_SHEET, sub_name="외주SUB(위트)")