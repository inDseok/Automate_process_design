from typing import List
from backend.models import SubTree, SubNode
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

# 업로드된 엑셀에서만 파싱 (로컬 경로 사용 안 함)
TARGET_SHEET = "1. SUB 단위 부품구성도(STD)"


def list_sub_names() -> List[str]:
    # 현재 구조에서는 업로드된 엑셀 기준으로만 동작
    return ["외주SUB(위트)"]


def extract_qty_from_text(text: str):
    if text is None:
        return None

    s = str(text).upper().replace(" ", "")
    m = re.search(r"(\d+(\.\d+)?)EA", s)
    if not m:
        return None

    try:
        return float(m.group(1))
    except:
        return None


def read_right_text(ws, r, c):
    texts = []
    col = c + 1
    max_col = ws.max_column

    while col <= max_col:
        cell = ws.cell(row=r, column=col)
        cell_coord = f"{get_column_letter(col)}{r}"

        merged_range = None
        for merged in ws.merged_cells.ranges:
            if cell_coord in merged:
                merged_range = merged
                break

        if merged_range:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
            if top_left not in [None, "", "부품명"]:
                texts.append(str(top_left).strip())
            col = merged_range.max_col + 1
            continue

        if cell.value not in [None, "", "부품명"]:
            texts.append(str(cell.value).strip())

        col += 1
        if texts:
            break

    unique = []
    for t in texts:
        if t not in unique:
            unique.append(t)

    return " ".join(unique).strip()


def find_row_with_label(ws, start_r, start_c, label):
    for offset in range(1, 10):
        row = start_r + offset
        col = start_c
        cell = ws.cell(row=row, column=col)

        if cell.value:
            s = str(cell.value).replace(" ", "").replace("\n", "").strip()
            if label in s:
                return offset

        coord = f"{get_column_letter(col)}{row}"
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                top_left = ws.cell(merged.min_row, merged.min_col)
                if top_left.value:
                    s2 = str(top_left.value).replace(" ", "").replace("\n", "").strip()
                    if label in s2:
                        return offset

    raise ValueError(f"Label '{label}' not found below row {start_r}")


def read_qty_robust(ws, start_row, start_col):
    for r in range(start_row, start_row + 6):
        for c in range(start_col, start_col + 12):
            coord = f"{get_column_letter(c)}{r}"
            value = ws.cell(row=r, column=c).value

            for merged in ws.merged_cells.ranges:
                if coord in merged:
                    value = ws.cell(merged.min_row, merged.min_col).value
                    break

            qty = extract_qty_from_text(value)
            if qty is not None:
                return qty

    return None


def read_vehicle(ws, r):
    for col in range(1, 200):
        cell = ws.cell(row=r, column=col)
        if cell.value == "양산처":
            val_cell = ws.cell(row=r, column=col + 1)
            return str(val_cell.value).strip() if val_cell.value is not None else None
    return None


def make_stable_id(sheet_name: str, row: int, col: int) -> str:
    return f"{sheet_name}:{row}:{col}"


def parse_block(ws, sheet_name: str, label_row: int, label_col: int):
    r = label_row
    c = label_col

    name = read_right_text(ws, r, c)

    row2_offset = find_row_with_label(ws, r, c, "양산처")
    row2 = r + row2_offset
    vehicle = read_vehicle(ws, row2)

    material_row_offset = find_row_with_label(ws, r, c, "재질")
    material = read_right_text(ws, r + material_row_offset, c)

    qty = read_qty_robust(ws, r, c)

    return {
        "id": make_stable_id(sheet_name, r, c),
        "name": name,
        "qty": qty,
        "vehicle": vehicle,
        "material": material,
        "row": r,
        "col": c,
    }


def build_tree_from_sheet(ws, sheet_name: str, sub_name: str) -> SubTree:
    boxes = []

    for row in ws.iter_rows(min_row=1, max_col=200, max_row=500):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if str(v).strip() == "부품명":
                box = parse_block(ws, sheet_name, cell.row, cell.column)
                boxes.append(box)

    boxes_sorted = sorted(boxes, key=lambda b: (b["row"], b["col"]))

    stack = []
    nodes: List[SubNode] = []

    for idx, box in enumerate(boxes_sorted):
        while stack and box["col"] <= stack[-1]["col"]:
            stack.pop()

        parent_id = stack[-1]["id"] if stack else None

        nodes.append(
            SubNode(
                id=box["id"],
                parent_id=parent_id,
                order=idx,
                type="PART",
                name=box["name"],
                vehicle=box["vehicle"],
                material=box["material"],
                qty=box["qty"],
            )
        )
        stack.append(box)

    return SubTree(sub_name=sub_name, nodes=nodes)


def parse_uploaded_excel(binary_data: bytes) -> SubTree:
    wb = load_workbook(BytesIO(binary_data), data_only=True)

    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"{TARGET_SHEET} 시트를 찾을 수 없습니다.")

    ws = wb[TARGET_SHEET]

    return build_tree_from_sheet(ws, TARGET_SHEET, sub_name="외주SUB(위트)")
